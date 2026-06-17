"""Buckholtz Results Log (RU) — Tom-Lawrence template, status in PERCENT.
Exact columns requested by user. Saves xlsx to Downloads.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = Path.home() / "Downloads" / "buckholtz_results_log_RU_v1.xlsx"

DARK = PatternFill("solid", fgColor="1F2937")
SECT = PatternFill("solid", fgColor="374151")
HEAD = PatternFill("solid", fgColor="4B5563")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
WHITE_B = Font(color="FFFFFF", bold=True, size=11)
WHITE_S = Font(color="FFFFFF", bold=True, size=10)
thin = Side(style="thin", color="9CA3AF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def fill_for(pct: int) -> PatternFill:
    if pct >= 70:
        return GREEN
    if pct >= 40:
        return YELLOW
    return RED


wb = Workbook()
ws = wb.active
ws.title = "Журнал результатов"

ws.merge_cells("A1:F1")
c = ws["A1"]
c.value = (
    "Проект: Buckholtz IDM/MULTING  |  Обновлено: 2026-06-17  |  "
    "Звонок TJB 2026-06-14: k_A и мост подтверждены (docs/117)  |  "
    "HARD: NOT_VALIDATION · NOT_REFUTATION · OUR_RECONSTRUCTION"
)
c.fill = DARK
c.font = WHITE_S
c.alignment = WRAP
ws.row_dimensions[1].height = 44

row = 3


def section(title: str) -> None:
    global row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cc = ws.cell(row=row, column=1, value=title)
    cc.fill = SECT
    cc.font = WHITE_B
    ws.row_dimensions[row].height = 22
    row += 1


def headers(cols: list[str]) -> None:
    global row
    for j, name in enumerate(cols, start=1):
        cc = ws.cell(row=row, column=j, value=name)
        cc.fill = HEAD
        cc.font = WHITE_S
        cc.alignment = WRAP
        cc.border = BORDER
    ws.row_dimensions[row].height = 32
    row += 1


def datarow(values: list, status_col: int | None = None) -> None:
    global row
    for j, val in enumerate(values, start=1):
        out = val
        if status_col and j == status_col and isinstance(val, int):
            out = f"{val}%"
        cc = ws.cell(row=row, column=j, value=out)
        cc.alignment = WRAP
        cc.border = BORDER
        if status_col and j == status_col and isinstance(val, int):
            cc.fill = fill_for(val)
            cc.alignment = CENTER
    ws.row_dimensions[row].height = 58
    row += 1


# ── SECTION 1 — Results Log (exact columns requested) ────────────────────────
section("СЕКЦИЯ 1 — ЖУРНАЛ РЕЗУЛЬТАТОВ")
headers(
    [
        "#",
        "Входные данные / Отправная точка",
        "Статус (0–100%)",
        "Выполненная работа: Как получен результат",
        "Результат / Формула / Доказательство",
        "Интерпретация: Что это даёт теории",
    ]
)

results = [
    (
        1,
        "C9: (4/3)(m_τ/m_e)¹² = α_EM/α_G (Eq.32)",
        95,
        "Вычисление с PDG 2024; α_G через массу электрона (не протона)",
        "Отклонение 0.17σ  [VERIFIED]",
        "Самое точное соотношение теории. Ключ: α_G через m_e. Эмпирическое — механизм не выведен.",
    ),
    (
        2,
        "C6: m_W²:m_Z²:m_H² = 7:9:17 (Eq.31)",
        85,
        "PDG 2024; расчёт σ (нормировка несёт точность m_W)",
        "Higgs 0.4σ ✅; Z 5.5σ ⚠️  [VERIFIED]",
        "Higgs-соотношение подтверждено. Z-соотношение остаётся открытым.",
    ),
    (
        3,
        "Размерности F_d, F_q (Eq.15–16)",
        90,
        "Размерный анализ дипольного и квадрупольного членов",
        "Размерность корректна  [VERIFIED]",
        "Формулы сил внутренне согласованы. Не вывод из лагранжиана.",
    ),
    (
        4,
        "Инфлатон из формулы бозонов (Eq.27–30)",
        70,
        "Подстановка в (N')² = Σ квадратов",
        "m_inflaton ≈ m_Z/3 ≈ 30.4 ГэВ",
        "Фальсифицируемое предсказание; пока не наблюдаемо.",
    ),
    (
        5,
        "Счёт изомеров 5:1 (IDM)",
        70,
        "Подсчёт: 5 тёмных изомеров / 1 обычный",
        "Согласуется с наблюдаемым (5+):1",
        "Внутренне последовательно. НЕ derivation — 5 механизмов отсутствуют.",
    ),
    (
        6,
        "k_A — физический смысл",
        85,
        "[TOM CONFIRMED 2026-06-14, docs/117]",
        "k_A = E_ICM/c² (в M☉) ~ 10⁸–10⁹",
        "Прояснено автором. Вопрос «что такое k_A» закрыт; остаётся β-rescaling.",
    ),
    (
        7,
        "Мост F_oP → H(z) (Appendix A.1)",
        40,
        "[TOM CONFIRMED 2026-06-14] мост делегирован мне",
        "phenomenological mapping; маршрут кинематический (сила→ä/a→H(z)), не ОТО",
        "Строю сам, как делегировал автор. ОТО-аналог — future work. НЕ доказывает MULTING.",
    ),
    (
        8,
        "Pearson r: 443 кластера (MCXC) + CC H(z) (Moresco+2022)",
        30,
        "r(H_MULT, H_CC) на реальных каталогах",
        "r=0.62 vs trivial=0.89  [VERIFIED-real]",
        "INCONCLUSIVE на flux-limited выборке — не различает MULTING и Malmquist bias.",
    ),
    (
        9,
        "23 кластера с XMM T_X (Rossetti+2024)",
        40,
        "r; r(logM,z)=0.094 (selection bias ≈ 0)",
        "r=0.70 vs trivial=0.91  [VERIFIED-real]",
        "Для решающего теста нужна volume-limited выборка с XMM T_X.",
    ),
    (
        10,
        "AIC/BIC: ΛCDM vs MULTING на CC H(z)",
        40,
        "Сравнение информационных критериев",
        "in-sample; ΔAIC=−16.65 на Table A1 (только IN-SAMPLE)",
        "Требуется out-of-sample тест после построения моста.",
    ),
]
for r_ in results:
    datarow(list(r_), status_col=3)

row += 1

# ── SECTION 2 — Key Insights ─────────────────────────────────────────────────
section("СЕКЦИЯ 2 — КЛЮЧЕВЫЕ ИНСАЙТЫ")
headers(
    [
        "#",
        "Наблюдение",
        "Статус (0–100%)",
        "Инсайт",
        "Формула / Проверяемое ядро",
        "",
    ]
)

insights = [
    (
        1,
        "AI-сервисы выбрали разные β",
        90,
        "β — фитируемые параметры без единого канонического значения",
        "ChatGPT 0.78/0.19 · Claude 4.5/18.0 · Gemini 4.25/8.10 → β_d ~5.8×, β_q ~95×",
        "",
    ),
    (
        2,
        "β-rescaling при физическом k_A",
        70,
        "β_d=4.5 (Table A1) калибровано под свободное β·radius (TJB: не ограничивать длинами)",
        "при k_A=E_ICM/c² нужен масштаб ~×10³ к Table A1",
        "",
    ),
    (
        3,
        "Монополь стабильно лучше полной формулы",
        60,
        "Дипольный член добавляет scatter, а не сигнал — при текущей нормировке",
        "r(монополь) > r(F_oP) во всех тестах",
        "",
    ),
    (
        4,
        "Знак дипольного члена",
        50,
        "CC-данные требуют положительный знак, MULTING — отрицательный",
        "CC: [+,+,+]  vs  MULTING: [+,−,+] — нужен screening-механизм?",
        "",
    ),
]
for i_ in insights:
    datarow(list(i_), status_col=3)

row += 1

# ── SECTION 3 — Questions for TJB ────────────────────────────────────────────
section("СЕКЦИЯ 3 — ВОПРОСЫ К Dr. BUCKHOLTZ (только то, что можете сказать вы)")
headers(["Q#", "Вопрос к Тому", "Зачем спрашиваю", "Что даёт ответ", "", ""])

questions = [
    (
        "Q1",
        "Есть ли целевой срок (журнал / следующая версия препринта), к которому полезно подготовить результаты? Сколько у меня времени и в каком темпе вам удобно?",
        "Чтобы расставить приоритеты под ваш темп.",
        "Срок есть → план фаз под дедлайн.  Срока нет → исследовательский темп без спешки.",
        "",
        "",
    ),
    (
        "Q2",
        "Что для вас сейчас ценнее: (a) воспроизводимый scaffold Table A1, (b) сравнение β между AI-сервисами, (c) фит на volume-limited выборке с XMM T_X?",
        "Чтобы сфокусироваться на одном артефакте, а не распыляться.",
        "(a) scaffold+provenance · (b) методологическая заметка · (c) решающий Pearson-r тест (ваш Step 4).",
        "",
        "",
    ),
]
for q_ in questions:
    datarow(list(q_))

widths = [5, 34, 14, 38, 38, 44]
for j, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + j)].width = w

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"SAVED: {OUT}")
